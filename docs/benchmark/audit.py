#!/usr/bin/env python3
"""Showdown v2 Auditor — checks both FCP and Raw xlsx files against the spec."""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DIR = Path(__file__).parent


def audit_file(path: str, label: str) -> dict:
    """Audit a single xlsx file. Returns {checks: [...], score: (pass, total)}."""
    wb = load_workbook(path, data_only=False)
    checks = []

    def check(category, name, passed, detail=""):
        checks.append({
            "category": category,
            "name": name,
            "passed": passed,
            "detail": detail
        })

    sheet_names = wb.sheetnames

    # ═══════════════════════════════════════════════════════
    # PHASE 1 CHECKS
    # ═══════════════════════════════════════════════════════

    # --- Structure ---
    expected_sheets_p1 = ["Executive Summary", "Fund Performance", "Holdings", "Portfolio Detail", "Cash Flows", "Scenario Analysis"]
    # After Phase 2, "Holdings" is renamed to "Portfolio Detail" and "Sector Allocation" is added
    has_portfolio_detail = "Portfolio Detail" in sheet_names
    has_holdings = "Holdings" in sheet_names
    holdings_name = "Portfolio Detail" if has_portfolio_detail else "Holdings"

    expected_final = ["Executive Summary", "Fund Performance", holdings_name, "Cash Flows", "Scenario Analysis", "Sector Allocation"]

    check("Structure", "Has Executive Summary sheet", "Executive Summary" in sheet_names)
    check("Structure", "Has Fund Performance sheet", "Fund Performance" in sheet_names)
    check("Structure", "Has Holdings/Portfolio Detail sheet", has_holdings or has_portfolio_detail,
          f"Found: {holdings_name}")
    check("Structure", "Has Cash Flows sheet", "Cash Flows" in sheet_names)
    check("Structure", "Has Scenario Analysis sheet", "Scenario Analysis" in sheet_names)
    check("Structure", "No default 'Sheet' tab", "Sheet" not in sheet_names,
          f"Sheets: {sheet_names}")

    # --- Executive Summary ---
    if "Executive Summary" in sheet_names:
        ws = wb["Executive Summary"]

        # Title merge
        merged = [str(m) for m in ws.merged_cells.ranges]
        check("ExecSummary", "Title merge A1:F1", any("A1:F1" in m for m in merged),
              f"Merges: {merged[:5]}")

        # Fund Overview data
        check("ExecSummary", "Fund Size = 500000000", ws["B5"].value == 500000000,
              f"B5={ws['B5'].value}")
        check("ExecSummary", "Vintage Year = 2021", ws["B6"].value == 2021,
              f"B6={ws['B6'].value}")
        check("ExecSummary", "Investment Period text", ws["B7"].value == "2021–2024" or ws["B7"].value == "2021-2024",
              f"B7={ws['B7'].value}")
        check("ExecSummary", "Invested Capital = 385000000", ws["B8"].value == 385000000,
              f"B8={ws['B8'].value}")
        check("ExecSummary", "Realized Value = 142000000", ws["B9"].value == 142000000,
              f"B9={ws['B9'].value}")
        check("ExecSummary", "Unrealized Value = 498000000", ws["B10"].value == 498000000,
              f"B10={ws['B10'].value}")
        # Total Value formula
        b11 = ws["B11"].value
        check("ExecSummary", "Total Value formula (=B9+B10)",
              isinstance(b11, str) and "B9" in b11 and "B10" in b11,
              f"B11={b11}")

        # Performance metrics
        check("ExecSummary", "Net IRR = 0.218", ws["B14"].value == 0.218,
              f"B14={ws['B14'].value}")
        check("ExecSummary", "Gross IRR = 0.267", ws["B15"].value == 0.267,
              f"B15={ws['B15'].value}")
        check("ExecSummary", "Net MOIC = 1.66", ws["B16"].value == 1.66,
              f"B16={ws['B16'].value}")
        # DPI formula
        b17 = ws["B17"].value
        check("ExecSummary", "DPI formula (=B9/B8)",
              isinstance(b17, str) and "B9" in str(b17) and "B8" in str(b17),
              f"B17={b17}")

        # Portfolio Snapshot
        check("ExecSummary", "Active Companies = 8", ws["B22"].value == 8,
              f"B22={ws['B22'].value}")
        check("ExecSummary", "Fully Exited = 2", ws["B23"].value == 2,
              f"B23={ws['B23'].value}")

        # Named ranges
        defined_names = {dn.name: str(dn.value) for dn in wb.defined_names.values()}
        check("ExecSummary", "Named range FundSize", "FundSize" in defined_names,
              f"Names: {list(defined_names.keys())[:10]}")
        check("ExecSummary", "Named range InvestedCapital", "InvestedCapital" in defined_names)
        check("ExecSummary", "Named range TotalValue", "TotalValue" in defined_names)

        # Protection
        check("ExecSummary", "Sheet protected", ws.protection.sheet,
              f"protection.sheet={ws.protection.sheet}")

    # --- Fund Performance ---
    if "Fund Performance" in sheet_names:
        ws = wb["Fund Performance"]

        # Title merge
        merged = [str(m) for m in ws.merged_cells.ranges]
        check("FundPerf", "Title merge A1:H1", any("A1:H1" in m or "A1:I1" in m for m in merged),
              f"Merges: {merged[:5]}")

        # Data values (row 3 = 2021)
        check("FundPerf", "2021 Contributions = -125000000", ws["B3"].value == -125000000,
              f"B3={ws['B3'].value}")
        check("FundPerf", "2022 Distributions = 12000000", ws["C4"].value == 12000000,
              f"C4={ws['C4'].value}")
        check("FundPerf", "2025 Net IRR = 0.218", ws["F7"].value == 0.218,
              f"F7={ws['F7'].value}")
        check("FundPerf", "2025 MOIC = 1.66", ws["G7"].value == 1.66,
              f"G7={ws['G7'].value}")

        # Net Cash Flow formulas
        d3 = ws["D3"].value
        check("FundPerf", "Net Cash Flow formula D3 (=B3+C3)",
              isinstance(d3, str) and "B3" in str(d3) and "C3" in str(d3),
              f"D3={d3}")

        # Total row formulas
        b8 = ws["B8"].value
        check("FundPerf", "Total Contributions formula (SUM)",
              isinstance(b8, str) and "SUM" in str(b8).upper(),
              f"B8={b8}")

        # Freeze pane
        check("FundPerf", "Freeze pane at A3", ws.freeze_panes == "A3",
              f"freeze={ws.freeze_panes}")

        # Charts
        chart_count = len(ws._charts)
        check("FundPerf", "Has 2 charts", chart_count >= 2,
              f"Charts: {chart_count}")

    # --- Holdings / Portfolio Detail ---
    if holdings_name in sheet_names:
        ws = wb[holdings_name]

        # Company data (row 3 = Helios Energy)
        check("Holdings", "Helios Energy in A3",
              ws["A3"].value and "Helios" in str(ws["A3"].value),
              f"A3={ws['A3'].value}")
        check("Holdings", "Helios Cost Basis = 65000000", ws["D3"].value == 65000000,
              f"D3={ws['D3'].value}")
        check("Holdings", "Helios Fair Value = 142000000", ws["E3"].value == 142000000,
              f"E3={ws['E3'].value}")

        # MOIC formula
        f3 = ws["F3"].value
        check("Holdings", "MOIC formula F3 (=E3/D3)",
              isinstance(f3, str) and "E3" in str(f3) and "D3" in str(f3),
              f"F3={f3}")

        # Written Off company
        check("Holdings", "Bolt Payments status = Written Off",
              ws["K11"].value == "Written Off",
              f"K11={ws['K11'].value}")

        # Apex Dynamics (Exited)
        check("Holdings", "Apex Dynamics status = Exited",
              ws["K12"].value == "Exited",
              f"K12={ws['K12'].value}")

        # Totals row formulas
        d13 = ws["D13"].value
        check("Holdings", "Totals Cost Basis formula (SUM)",
              isinstance(d13, str) and "SUM" in str(d13).upper(),
              f"D13={d13}")

        # Freeze pane
        check("Holdings", "Freeze pane at A3", ws.freeze_panes == "A3",
              f"freeze={ws.freeze_panes}")

        # Filter
        check("Holdings", "AutoFilter enabled", ws.auto_filter.ref is not None,
              f"filter={ws.auto_filter.ref}")

        # Chart
        chart_count = len(ws._charts)
        check("Holdings", "Has chart (bubble/scatter)", chart_count >= 1,
              f"Charts: {chart_count}")

    # --- Cash Flows ---
    if "Cash Flows" in sheet_names:
        ws = wb["Cash Flows"]

        # Title merge
        merged = [str(m) for m in ws.merged_cells.ranges]
        check("CashFlows", "Title merge A1:F1 or A1:G1",
              any("A1:F1" in m or "A1:G1" in m for m in merged),
              f"Merges: {merged[:5]}")

        # Data values
        check("CashFlows", "Q1 2021 Contributions = -50000000", ws["B3"].value == -50000000,
              f"B3={ws['B3'].value}")
        check("CashFlows", "Q1 2021 NAV = 48000000", ws["F3"].value == 48000000,
              f"F3={ws['F3'].value}")
        check("CashFlows", "Q4 2025 NAV = 640000000", ws["F22"].value == 640000000,
              f"F22={ws['F22'].value}")

        # Net Cash Flow formula
        d3 = ws["D3"].value
        check("CashFlows", "Net Cash Flow formula D3 (=B3+C3)",
              isinstance(d3, str) and "B3" in str(d3) and "C3" in str(d3),
              f"D3={d3}")

        # Cumulative Cash Flow formula
        e4 = ws["E4"].value
        check("CashFlows", "Cumulative formula E4 (=E3+D4)",
              isinstance(e4, str) and "E3" in str(e4) and "D4" in str(e4),
              f"E4={e4}")

        # Total row
        b23 = ws["B23"].value
        check("CashFlows", "Total Contributions formula (SUM)",
              isinstance(b23, str) and "SUM" in str(b23).upper(),
              f"B23={b23}")

        # Named range LatestNAV
        check("CashFlows", "Named range LatestNAV", "LatestNAV" in defined_names,
              f"Names: {list(defined_names.keys())[:10]}")

        # Freeze
        check("CashFlows", "Freeze pane at A3", ws.freeze_panes == "A3",
              f"freeze={ws.freeze_panes}")

        # Charts
        chart_count = len(ws._charts)
        check("CashFlows", "Has 2 charts", chart_count >= 2,
              f"Charts: {chart_count}")

    # --- Scenario Analysis ---
    if "Scenario Analysis" in sheet_names:
        ws = wb["Scenario Analysis"]

        # Assumptions data
        check("ScenarioAnalysis", "Revenue Growth Base = 0.25", ws["B5"].value == 0.25,
              f"B5={ws['B5'].value}")
        check("ScenarioAnalysis", "Revenue Growth Bull = 0.40", ws["C5"].value == 0.40 or ws["C5"].value == 0.4,
              f"C5={ws['C5'].value}")
        check("ScenarioAnalysis", "Loss Rate Bear = 0.15", ws["D9"].value == 0.15,
              f"D9={ws['D9'].value}")

        # Cross-sheet formula (Portfolio Revenue)
        b13 = ws["B13"].value
        check("ScenarioAnalysis", "Portfolio Revenue cross-sheet formula",
              isinstance(b13, str) and ("Holdings" in str(b13) or "Portfolio Detail" in str(b13)),
              f"B13={b13}")

        # Implied MOIC cross-sheet formula
        b18 = ws["B18"].value
        check("ScenarioAnalysis", "Implied MOIC cross-sheet formula",
              isinstance(b18, str) and "Executive Summary" in str(b18),
              f"B18={b18}")

        # Chart
        chart_count = len(ws._charts)
        check("ScenarioAnalysis", "Has MOIC comparison chart", chart_count >= 1,
              f"Charts: {chart_count}")

    # ═══════════════════════════════════════════════════════
    # PHASE 2 CHECKS
    # ═══════════════════════════════════════════════════════

    # --- Cash Flows Total Return column ---
    if "Cash Flows" in sheet_names:
        ws = wb["Cash Flows"]
        g2 = ws["G2"].value
        check("Phase2-CF", "Total Return header in G2",
              g2 and "Total Return" in str(g2),
              f"G2={g2}")

        g3 = ws["G3"].value
        check("Phase2-CF", "Total Return formula G3 (Cumulative + NAV)",
              isinstance(g3, str) and ("E3" in str(g3) or "F3" in str(g3)),
              f"G3={g3}")

        g22 = ws["G22"].value
        check("Phase2-CF", "Total Return has value in G22",
              g22 is not None,
              f"G22={g22}")

        # Total row: should be =G22, not SUM
        g23 = ws["G23"].value
        check("Phase2-CF", "Total Return total = G22 ref (not SUM)",
              isinstance(g23, str) and "G22" in str(g23),
              f"G23={g23}")

    # --- Holdings renamed to Portfolio Detail ---
    check("Phase2-Holdings", "Sheet renamed to 'Portfolio Detail'", has_portfolio_detail,
          f"Found: {'Portfolio Detail' if has_portfolio_detail else 'Holdings'}")

    # --- Investment Thesis column (L) ---
    if holdings_name in sheet_names:
        ws = wb[holdings_name]
        l2 = ws["L2"].value
        check("Phase2-Holdings", "Investment Thesis header in L2",
              l2 and "Thesis" in str(l2),
              f"L2={l2}")

        l3 = ws["L3"].value
        check("Phase2-Holdings", "Helios thesis = 'Grid-scale battery storage leader'",
              l3 and "battery" in str(l3).lower(),
              f"L3={l3}")

        l6 = ws["L6"].value
        check("Phase2-Holdings", "Prism thesis = 'AI-native business intelligence'",
              l6 and "AI" in str(l6),
              f"L6={l6}")

    # --- Watchlist column (M) ---
    if holdings_name in sheet_names:
        ws = wb[holdings_name]
        m2 = ws["M2"].value
        check("Phase2-Holdings", "Watchlist header in M2",
              m2 and "Watchlist" in str(m2),
              f"M2={m2}")

        m3 = ws["M3"].value
        check("Phase2-Holdings", "Helios Watchlist = Green",
              m3 and str(m3).strip().lower() == "green",
              f"M3={m3}")

        m4 = ws["M4"].value
        check("Phase2-Holdings", "NovaBio Watchlist = Red",
              m4 and str(m4).strip().lower() == "red",
              f"M4={m4}")

        m7 = ws["M7"].value
        check("Phase2-Holdings", "Stratos Watchlist = Yellow",
              m7 and str(m7).strip().lower() == "yellow",
              f"M7={m7}")

    # --- Sector Allocation sheet ---
    check("Phase2-Sector", "Has Sector Allocation sheet", "Sector Allocation" in sheet_names,
          f"Sheets: {sheet_names}")

    if "Sector Allocation" in sheet_names:
        ws = wb["Sector Allocation"]

        # Title
        a1 = ws["A1"].value
        check("Phase2-Sector", "Title contains 'Sector Allocation'",
              a1 and "Sector" in str(a1),
              f"A1={a1}")

        # Header row
        a2 = ws["A2"].value
        check("Phase2-Sector", "Header row exists (A2=Sector)",
              a2 and "Sector" in str(a2),
              f"A2={a2}")

        # Software sector (should be first or present)
        found_software = False
        for row in range(3, 12):
            val = ws[f"A{row}"].value
            if val and "Software" in str(val):
                found_software = True
                # Check company count
                b_val = ws[f"B{row}"].value
                check("Phase2-Sector", "Software companies count = 2",
                      b_val == 2,
                      f"B{row}={b_val}")
                break
        check("Phase2-Sector", "Software sector present", found_software)

        # Chart
        chart_count = len(ws._charts)
        check("Phase2-Sector", "Has doughnut/pie chart", chart_count >= 1,
              f"Charts: {chart_count}")

    # ═══════════════════════════════════════════════════════
    # PHASE 3 CHECKS
    # ═══════════════════════════════════════════════════════

    # --- Navigation hyperlinks ---
    if "Executive Summary" in sheet_names:
        ws = wb["Executive Summary"]
        a26 = ws["A26"].value
        check("Phase3", "Navigation section header A26",
              a26 and "Navigation" in str(a26),
              f"A26={a26}")

        # Check for hyperlinks in A27-A31
        has_links = False
        for row in range(27, 32):
            cell = ws[f"A{row}"]
            if cell.hyperlink is not None:
                has_links = True
                break
        check("Phase3", "Has navigation hyperlinks (A27-A31)", has_links)

    # --- Data bars on Fund Performance distributions ---
    if "Fund Performance" in sheet_names:
        ws = wb["Fund Performance"]
        cf_rules = ws.conditional_formatting._cf_rules if hasattr(ws.conditional_formatting, '_cf_rules') else list(ws.conditional_formatting)
        has_data_bars = False
        for rule in cf_rules:
            rule_str = str(rule)
            if "dataBar" in rule_str.lower() or "DataBar" in str(type(rule)):
                has_data_bars = True
                break
            # Check rule object
            if hasattr(rule, 'rule'):
                if hasattr(rule.rule, 'dataBar') and rule.rule.dataBar is not None:
                    has_data_bars = True
                    break
        # Also check via the conditional formatting list
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.dataBar is not None:
                    has_data_bars = True
                    break
        check("Phase3", "Data bars on Fund Performance distributions", has_data_bars)

    # --- Page footers ---
    footer_count = 0
    for sn in sheet_names:
        ws = wb[sn]
        footer = ws.oddFooter
        if footer and ("Meridian" in str(footer) or "Confidential" in str(footer)):
            footer_count += 1
    check("Phase3", "Page footers set on sheets",
          footer_count >= 3,
          f"Sheets with footers: {footer_count}/{len(sheet_names)}")

    # --- Print title rows ---
    if holdings_name in sheet_names:
        ws = wb[holdings_name]
        check("Phase3", f"Print titles on {holdings_name}",
              ws.print_title_rows is not None,
              f"print_title_rows={ws.print_title_rows}")

    if "Cash Flows" in sheet_names:
        ws = wb["Cash Flows"]
        check("Phase3", "Print titles on Cash Flows",
              ws.print_title_rows is not None,
              f"print_title_rows={ws.print_title_rows}")

    # --- LatestMOIC named range ---
    check("Phase3", "Named range LatestMOIC", "LatestMOIC" in defined_names,
          f"Names: {list(defined_names.keys())}")

    # ═══════════════════════════════════════════════════════
    # Summary
    # ═══════════════════════════════════════════════════════
    passed = sum(1 for c in checks if c["passed"])
    total = len(checks)

    return {
        "label": label,
        "checks": checks,
        "passed": passed,
        "total": total,
    }


def print_results(result):
    """Print audit results."""
    print(f"\n{'=' * 65}")
    print(f"AUDIT: {result['label']}")
    print(f"{'=' * 65}")

    current_cat = None
    for c in result["checks"]:
        if c["category"] != current_cat:
            current_cat = c["category"]
            print(f"\n  [{current_cat}]")

        icon = "PASS" if c["passed"] else "FAIL"
        detail = f" — {c['detail']}" if c["detail"] and not c["passed"] else ""
        print(f"    [{icon}] {c['name']}{detail}")

    print(f"\n  Score: {result['passed']}/{result['total']} ({100*result['passed']/result['total']:.0f}%)")


def main():
    fcp_path = DIR / "v2-fcp-result.xlsx"
    raw_path = DIR / "v2-raw-result.xlsx"

    results = {}

    if fcp_path.exists():
        results["fcp"] = audit_file(str(fcp_path), "FCP (fcp-sheets DSL)")
        print_results(results["fcp"])
    else:
        print(f"SKIP: {fcp_path} not found")

    if raw_path.exists():
        results["raw"] = audit_file(str(raw_path), "Raw (openpyxl Python)")
        print_results(results["raw"])
    else:
        print(f"SKIP: {raw_path} not found")

    # Comparison
    if len(results) == 2:
        print(f"\n{'=' * 65}")
        print("COMPARISON SUMMARY")
        print(f"{'=' * 65}")
        print(f"  {'Check':<45} {'FCP':>5} {'Raw':>5}")
        print(f"  {'─' * 55}")

        fcp_checks = {c["name"]: c["passed"] for c in results["fcp"]["checks"]}
        raw_checks = {c["name"]: c["passed"] for c in results["raw"]["checks"]}

        all_names = list(dict.fromkeys(
            [c["name"] for c in results["fcp"]["checks"]] +
            [c["name"] for c in results["raw"]["checks"]]
        ))

        disagree = 0
        for name in all_names:
            fcp_p = fcp_checks.get(name, False)
            raw_p = raw_checks.get(name, False)
            fcp_icon = "PASS" if fcp_p else "FAIL"
            raw_icon = "PASS" if raw_p else "FAIL"
            marker = " <<" if fcp_p != raw_p else ""
            print(f"  {name:<45} {fcp_icon:>5} {raw_icon:>5}{marker}")
            if fcp_p != raw_p:
                disagree += 1

        fcp_r = results["fcp"]
        raw_r = results["raw"]
        print(f"\n  TOTAL: FCP {fcp_r['passed']}/{fcp_r['total']} vs Raw {raw_r['passed']}/{raw_r['total']}")
        if disagree:
            print(f"  Disagreements: {disagree} checks (marked <<)")

    # Save JSON
    output_path = DIR / "v2-audit-results.json"
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\nResults saved to {output_path}")


if __name__ == "__main__":
    main()
