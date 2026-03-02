"""Tests for miscellaneous operations — name, image, link, comment, protect, lock, page-setup."""

from __future__ import annotations

import os
import tempfile

import pytest
from fcp_core import ParsedOp

def _has_pillow() -> bool:
    try:
        import PIL  # noqa: F401
        return True
    except ImportError:
        return False


from fcp_sheets.server.ops_misc import (
    op_comment,
    op_image,
    op_link,
    op_lock,
    op_name,
    op_page_setup,
    op_protect,
    op_unlock,
    op_unprotect,
)
from fcp_sheets.server.resolvers import SheetsOpContext


# ── name ────────────────────────────────────────────────────────────────

class TestNameDefine:
    def test_define_named_range(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["define", "SalesData"],
            params={"range": "$A$1:$D$10"}, raw="name define SalesData range:$A$1:$D$10",
        )
        result = op_name(op, ctx)
        assert result.success
        assert "SalesData" in result.message
        # Verify defined name exists in workbook
        assert "SalesData" in ctx.wb.defined_names

    def test_define_with_scope(self, ctx: SheetsOpContext):
        ctx.wb.create_sheet("Revenue")
        op = ParsedOp(
            verb="name", positionals=["define", "TotalRev"],
            params={"range": "$A$1:$A$100", "scope": "Revenue"},
            raw="name define TotalRev range:$A$1:$A$100 scope:Revenue",
        )
        result = op_name(op, ctx)
        assert result.success
        assert "Revenue" in result.message

    def test_define_missing_range(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["define", "NoRange"],
            raw="name define NoRange",
        )
        result = op_name(op, ctx)
        assert not result.success

    def test_define_missing_name(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["define"],
            raw="name define",
        )
        result = op_name(op, ctx)
        assert not result.success

    def test_define_scope_nonexistent_sheet(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["define", "Test"],
            params={"range": "$A$1", "scope": "Nonexistent"},
            raw="name define Test range:$A$1 scope:Nonexistent",
        )
        result = op_name(op, ctx)
        assert not result.success


class TestNameRemove:
    def test_remove_named_range(self, ctx: SheetsOpContext):
        # First define
        from openpyxl.workbook.defined_name import DefinedName
        defn = DefinedName("TestName", attr_text="'Sheet1'!$A$1:$B$5")
        ctx.wb.defined_names.add(defn)

        op = ParsedOp(
            verb="name", positionals=["remove", "TestName"],
            raw="name remove TestName",
        )
        result = op_name(op, ctx)
        assert result.success

    def test_remove_nonexistent(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["remove", "NoSuchName"],
            raw="name remove NoSuchName",
        )
        result = op_name(op, ctx)
        assert not result.success

    def test_name_unknown_action(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="name", positionals=["destroy", "Something"],
            raw="name destroy Something",
        )
        result = op_name(op, ctx)
        assert not result.success

    def test_name_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="name", positionals=[], raw="name")
        result = op_name(op, ctx)
        assert not result.success


# ── image ───────────────────────────────────────────────────────────────

class TestImage:
    def test_image_file_not_found(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="image", positionals=["A1"],
            params={"path": "/nonexistent/image.png"},
            raw="image A1 path:/nonexistent/image.png",
        )
        result = op_image(op, ctx)
        assert not result.success
        assert "not found" in result.message.lower()

    def test_image_missing_path(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="image", positionals=["A1"], raw="image A1")
        result = op_image(op, ctx)
        assert not result.success

    def test_image_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="image", positionals=[], raw="image")
        result = op_image(op, ctx)
        assert not result.success

    def test_image_invalid_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="image", positionals=["INVALID"],
            params={"path": "/tmp/test.png"},
            raw="image INVALID path:/tmp/test.png",
        )
        result = op_image(op, ctx)
        assert not result.success

    @pytest.mark.skipif(
        not _has_pillow(), reason="Pillow not installed"
    )
    def test_image_with_valid_file(self, ctx: SheetsOpContext):
        """Test image insertion with an actual temporary image file."""
        # Create a minimal 1x1 PNG file
        import struct
        import zlib

        def create_minimal_png() -> bytes:
            signature = b'\x89PNG\r\n\x1a\n'
            # IHDR chunk
            ihdr_data = struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
            ihdr_crc = zlib.crc32(b'IHDR' + ihdr_data)
            ihdr = struct.pack('>I', 13) + b'IHDR' + ihdr_data + struct.pack('>I', ihdr_crc & 0xffffffff)
            # IDAT chunk
            raw_data = b'\x00\xff\x00\x00'  # filter byte + RGB
            compressed = zlib.compress(raw_data)
            idat_crc = zlib.crc32(b'IDAT' + compressed)
            idat = struct.pack('>I', len(compressed)) + b'IDAT' + compressed + struct.pack('>I', idat_crc & 0xffffffff)
            # IEND chunk
            iend_crc = zlib.crc32(b'IEND')
            iend = struct.pack('>I', 0) + b'IEND' + struct.pack('>I', iend_crc & 0xffffffff)
            return signature + ihdr + idat + iend

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            f.write(create_minimal_png())
            tmp_path = f.name

        try:
            op = ParsedOp(
                verb="image", positionals=["A1"],
                params={"path": tmp_path},
                raw=f"image A1 path:{tmp_path}",
            )
            result = op_image(op, ctx)
            assert result.success
        finally:
            os.unlink(tmp_path)

    @pytest.mark.skipif(
        not _has_pillow(), reason="Pillow not installed"
    )
    def test_image_invalid_size_format(self, ctx: SheetsOpContext):
        """Test that invalid size format is rejected."""
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            # Write minimal PNG content
            import struct, zlib
            sig = b'\x89PNG\r\n\x1a\n'
            ihdr_data = struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
            ihdr_crc = zlib.crc32(b'IHDR' + ihdr_data)
            ihdr = struct.pack('>I', 13) + b'IHDR' + ihdr_data + struct.pack('>I', ihdr_crc & 0xffffffff)
            raw_data = b'\x00\xff\x00\x00'
            compressed = zlib.compress(raw_data)
            idat_crc = zlib.crc32(b'IDAT' + compressed)
            idat = struct.pack('>I', len(compressed)) + b'IDAT' + compressed + struct.pack('>I', idat_crc & 0xffffffff)
            iend_crc = zlib.crc32(b'IEND')
            iend = struct.pack('>I', 0) + b'IEND' + struct.pack('>I', iend_crc & 0xffffffff)
            f.write(sig + ihdr + idat + iend)
            tmp_path = f.name

        try:
            op = ParsedOp(
                verb="image", positionals=["A1"],
                params={"path": tmp_path, "size": "bad"},
                raw=f"image A1 path:{tmp_path} size:bad",
            )
            result = op_image(op, ctx)
            assert not result.success
            assert "size" in result.message.lower()
        finally:
            os.unlink(tmp_path)


# ── link ────────────────────────────────────────────────────────────────

class TestLink:
    def test_link_url(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="link", positionals=["A1"],
            params={"url": "https://example.com"},
            raw="link A1 url:https://example.com",
        )
        result = op_link(op, ctx)
        assert result.success
        cell = ctx.active_sheet.cell(row=1, column=1)
        assert cell.hyperlink is not None
        assert cell.value == "https://example.com"

    def test_link_url_with_text(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="link", positionals=["B2"],
            params={"url": "https://example.com", "text": "Click here"},
            raw='link B2 url:https://example.com text:"Click here"',
        )
        result = op_link(op, ctx)
        assert result.success
        cell = ctx.active_sheet.cell(row=2, column=2)
        assert cell.value == "Click here"

    def test_link_internal_sheet(self, ctx: SheetsOpContext):
        ctx.wb.create_sheet("Data")
        op = ParsedOp(
            verb="link", positionals=["A1"],
            params={"sheet": "Data!A1"},
            raw="link A1 sheet:Data!A1",
        )
        result = op_link(op, ctx)
        assert result.success
        cell = ctx.active_sheet.cell(row=1, column=1)
        assert cell.hyperlink is not None

    def test_link_internal_sheet_no_cell(self, ctx: SheetsOpContext):
        """Link to a sheet without specifying a cell defaults to A1."""
        ctx.wb.create_sheet("Summary")
        op = ParsedOp(
            verb="link", positionals=["A1"],
            params={"sheet": "Summary"},
            raw="link A1 sheet:Summary",
        )
        result = op_link(op, ctx)
        assert result.success

    def test_link_internal_nonexistent_sheet(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="link", positionals=["A1"],
            params={"sheet": "Nonexistent!A1"},
            raw="link A1 sheet:Nonexistent!A1",
        )
        result = op_link(op, ctx)
        assert not result.success

    def test_link_off(self, ctx: SheetsOpContext):
        # First add a link
        cell = ctx.active_sheet.cell(row=1, column=1)
        cell.hyperlink = "https://example.com"
        cell.value = "Link"

        op = ParsedOp(verb="link", positionals=["off", "A1"], raw="link off A1")
        result = op_link(op, ctx)
        assert result.success
        assert ctx.active_sheet.cell(row=1, column=1).hyperlink is None

    def test_link_off_missing_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="link", positionals=["off"], raw="link off")
        result = op_link(op, ctx)
        assert not result.success

    def test_link_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="link", positionals=[], raw="link")
        result = op_link(op, ctx)
        assert not result.success

    def test_link_missing_url_and_sheet(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="link", positionals=["A1"], raw="link A1")
        result = op_link(op, ctx)
        assert not result.success


# ── comment ─────────────────────────────────────────────────────────────

class TestComment:
    def test_add_comment(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="comment", positionals=["A1", "This is a note"],
            raw='comment A1 "This is a note"',
        )
        result = op_comment(op, ctx)
        assert result.success
        cell = ctx.active_sheet.cell(row=1, column=1)
        assert cell.comment is not None
        assert cell.comment.text == "This is a note"

    def test_remove_comment(self, ctx: SheetsOpContext):
        from openpyxl.comments import Comment
        cell = ctx.active_sheet.cell(row=1, column=1)
        cell.comment = Comment("Old note", "Author")

        op = ParsedOp(verb="comment", positionals=["off", "A1"], raw="comment off A1")
        result = op_comment(op, ctx)
        assert result.success
        assert ctx.active_sheet.cell(row=1, column=1).comment is None

    def test_comment_off_missing_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="comment", positionals=["off"], raw="comment off")
        result = op_comment(op, ctx)
        assert not result.success

    def test_comment_missing_text(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="comment", positionals=["A1"], raw="comment A1")
        result = op_comment(op, ctx)
        assert not result.success

    def test_comment_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="comment", positionals=[], raw="comment")
        result = op_comment(op, ctx)
        assert not result.success

    def test_comment_invalid_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="comment", positionals=["INVALID", "Note"],
            raw='comment INVALID "Note"',
        )
        result = op_comment(op, ctx)
        assert not result.success


# ── protect / unprotect ─────────────────────────────────────────────────

class TestProtect:
    def test_protect(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="protect", positionals=[], raw="protect")
        result = op_protect(op, ctx)
        assert result.success
        assert ctx.active_sheet.protection.sheet is True

    def test_protect_with_password(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="protect", positionals=[],
            params={"password": "secret123"},
            raw="protect password:secret123",
        )
        result = op_protect(op, ctx)
        assert result.success
        assert ctx.active_sheet.protection.sheet is True
        assert "password" in result.message.lower()

    def test_unprotect(self, ctx: SheetsOpContext):
        ctx.active_sheet.protection.sheet = True
        op = ParsedOp(verb="unprotect", positionals=[], raw="unprotect")
        result = op_unprotect(op, ctx)
        assert result.success
        assert ctx.active_sheet.protection.sheet is False


# ── lock / unlock ───────────────────────────────────────────────────────

class TestLockUnlock:
    def test_lock_range(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="lock", positionals=["A1:B2"], raw="lock A1:B2")
        result = op_lock(op, ctx)
        assert result.success
        for r in range(1, 3):
            for c in range(1, 3):
                assert ctx.active_sheet.cell(row=r, column=c).protection.locked is True

    def test_lock_single_cell(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="lock", positionals=["C3"], raw="lock C3")
        result = op_lock(op, ctx)
        assert result.success
        assert ctx.active_sheet.cell(row=3, column=3).protection.locked is True

    def test_unlock_range(self, ctx: SheetsOpContext):
        # Lock first
        from openpyxl.styles import Protection
        for r in range(1, 3):
            for c in range(1, 3):
                ctx.active_sheet.cell(row=r, column=c).protection = Protection(locked=True)

        op = ParsedOp(verb="unlock", positionals=["A1:B2"], raw="unlock A1:B2")
        result = op_unlock(op, ctx)
        assert result.success
        for r in range(1, 3):
            for c in range(1, 3):
                assert ctx.active_sheet.cell(row=r, column=c).protection.locked is False

    def test_lock_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="lock", positionals=[], raw="lock")
        result = op_lock(op, ctx)
        assert not result.success

    def test_unlock_no_args(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="unlock", positionals=[], raw="unlock")
        result = op_unlock(op, ctx)
        assert not result.success

    def test_lock_invalid_range(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="lock", positionals=["INVALID"], raw="lock INVALID")
        result = op_lock(op, ctx)
        assert not result.success


# ── page-setup ──────────────────────────────────────────────────────────

class TestPageSetup:
    def test_orientation_landscape(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"orient": "landscape"},
            raw="page-setup orient:landscape",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.page_setup.orientation == "landscape"

    def test_orientation_portrait(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"orient": "portrait"},
            raw="page-setup orient:portrait",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.page_setup.orientation == "portrait"

    def test_invalid_orientation(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"orient": "sideways"},
            raw="page-setup orient:sideways",
        )
        result = op_page_setup(op, ctx)
        assert not result.success

    def test_paper_size(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"paper": "a4"},
            raw="page-setup paper:a4",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.page_setup.paperSize == 9  # A4

    def test_unknown_paper_size(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"paper": "napkin"},
            raw="page-setup paper:napkin",
        )
        result = op_page_setup(op, ctx)
        assert not result.success

    def test_margins(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"margins": "1.0,0.75,1.0,0.75"},
            raw="page-setup margins:1.0,0.75,1.0,0.75",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        margins = ctx.active_sheet.page_margins
        assert margins.top == 1.0
        assert margins.right == 0.75
        assert margins.bottom == 1.0
        assert margins.left == 0.75

    def test_margins_wrong_count(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"margins": "1.0,0.75"},
            raw="page-setup margins:1.0,0.75",
        )
        result = op_page_setup(op, ctx)
        assert not result.success

    def test_print_area(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"print-area": "A1:D20"},
            raw="page-setup print-area:A1:D20",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        # openpyxl normalizes print_area to include sheet name and absolute refs
        assert "A" in ctx.active_sheet.print_area
        assert "D" in ctx.active_sheet.print_area
        assert "20" in ctx.active_sheet.print_area

    def test_fit_to_page(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"fit-width": "1", "fit-height": "0"},
            raw="page-setup fit-width:1 fit-height:0",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.page_setup.fitToWidth == 1
        assert ctx.active_sheet.page_setup.fitToHeight == 0

    def test_gridlines_flag(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=["gridlines"],
            raw="page-setup gridlines",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.print_options.gridLines is True

    def test_center_flags(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=["center-h", "center-v"],
            raw="page-setup center-h center-v",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.print_options.horizontalCentered is True
        assert ctx.active_sheet.print_options.verticalCentered is True

    def test_no_changes(self, ctx: SheetsOpContext):
        op = ParsedOp(verb="page-setup", positionals=[], raw="page-setup")
        result = op_page_setup(op, ctx)
        assert result.success
        assert "no" in result.message.lower()

    def test_multiple_settings(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=["gridlines"],
            params={"orient": "landscape", "paper": "letter", "print-area": "A1:Z50"},
            raw="page-setup orient:landscape paper:letter print-area:A1:Z50 gridlines",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        ws = ctx.active_sheet
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.paperSize == 1
        assert "A" in ws.print_area and "Z" in ws.print_area and "50" in ws.print_area
        assert ws.print_options.gridLines is True

    def test_header_footer(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"header": "My Report", "footer": "Page &P"},
            raw='page-setup header:"My Report" footer:"Page &P"',
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.oddHeader.center.text == "My Report"
        assert ctx.active_sheet.oddFooter.center.text == "Page &P"

    def test_print_title_rows(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"print-title-rows": "1:2"},
            raw="page-setup print-title-rows:1:2",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.print_title_rows == "$1:$2"
        assert "print-title-rows" in result.message

    def test_print_title_cols(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"print-title-cols": "A:B"},
            raw="page-setup print-title-cols:A:B",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.print_title_cols == "$A:$B"
        assert "print-title-cols" in result.message

    def test_print_title_rows_and_cols(self, ctx: SheetsOpContext):
        op = ParsedOp(
            verb="page-setup", positionals=[],
            params={"print-title-rows": "1:1", "print-title-cols": "A:A"},
            raw="page-setup print-title-rows:1:1 print-title-cols:A:A",
        )
        result = op_page_setup(op, ctx)
        assert result.success
        assert ctx.active_sheet.print_title_rows == "$1:$1"
        assert ctx.active_sheet.print_title_cols == "$A:$A"
