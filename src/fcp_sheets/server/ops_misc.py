"""Miscellaneous operation handlers — name, image, link, comment, protect, page-setup."""

from __future__ import annotations

import os

from openpyxl.comments import Comment
from openpyxl.styles import Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins, PrintOptions

from fcp_core import OpResult, ParsedOp

from fcp_sheets.model.refs import (
    RangeRef,
    col_to_index,
    index_to_col,
    parse_cell_ref,
    parse_range_ref,
)
from fcp_sheets.server.resolvers import SheetsOpContext


def op_name(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Define or remove named ranges.

    Syntax:
      name define NAME range:RANGE [scope:SHEET]
      name remove NAME
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: name define|remove NAME [range:RANGE]")

    action = op.positionals[0].lower()

    if action == "define":
        if len(op.positionals) < 2:
            return OpResult(success=False, message="Usage: name define NAME range:RANGE [scope:SHEET]")

        name = op.positionals[1]
        range_str = op.params.get("range")
        if not range_str:
            return OpResult(success=False, message="Missing 'range' parameter for name define")

        scope_sheet = op.params.get("scope")

        # Build the attr_text for the defined name
        if scope_sheet:
            if scope_sheet not in ctx.wb.sheetnames:
                return OpResult(success=False, message=f"Sheet '{scope_sheet}' not found")
            attr_text = f"'{scope_sheet}'!{range_str}"
        elif "!" in range_str:
            # Cross-sheet reference — already includes sheet name
            sheet_part, _, cell_part = range_str.partition("!")
            sheet_part = sheet_part.strip("'\"")
            if sheet_part not in ctx.wb.sheetnames:
                return OpResult(
                    success=False,
                    message=f"Sheet '{sheet_part}' not found in range '{range_str}'",
                )
            attr_text = f"'{sheet_part}'!{cell_part}"
        else:
            # Use active sheet
            sheet_name = ctx.active_sheet.title
            attr_text = f"'{sheet_name}'!{range_str}"

        defn = DefinedName(name, attr_text=attr_text)

        # If scope is specified, add to sheet-scoped names
        if scope_sheet:
            sheet_idx = ctx.wb.sheetnames.index(scope_sheet)
            defn.localSheetId = sheet_idx

        ctx.wb.defined_names.add(defn)

        return OpResult(success=True, message=f"Defined name '{name}' = {attr_text}", prefix="+")

    elif action == "remove":
        if len(op.positionals) < 2:
            return OpResult(success=False, message="Usage: name remove NAME")

        name = op.positionals[1]

        try:
            del ctx.wb.defined_names[name]
        except KeyError:
            return OpResult(success=False, message=f"Named range '{name}' not found")

        return OpResult(success=True, message=f"Removed named range '{name}'", prefix="-")

    else:
        return OpResult(success=False, message=f"Unknown name action: {action!r}. Use 'define' or 'remove'")


def op_image(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Insert image at cell.

    Syntax: image CELL path:PATH [size:WxH]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: image CELL path:PATH [size:WxH]")

    cell_str = op.positionals[0]
    path = op.params.get("path")
    if not path:
        return OpResult(success=False, message="Missing 'path' parameter. Usage: image CELL path:PATH")

    # Validate cell reference
    cell_ref = parse_cell_ref(cell_str)
    if not cell_ref:
        return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

    # Check file exists
    if not os.path.isfile(path):
        return OpResult(success=False, message=f"Image file not found: {path!r}")

    from openpyxl.drawing.image import Image

    try:
        img = Image(path)
    except ImportError:
        return OpResult(success=False, message="Pillow is required for image support. Install with: pip install Pillow")

    # Apply size if specified
    size_str = op.params.get("size")
    if size_str:
        parts = size_str.lower().split("x")
        if len(parts) != 2:
            return OpResult(success=False, message=f"Invalid size format: {size_str!r}. Use WxH (e.g. 200x100)")
        try:
            img.width = int(parts[0])
            img.height = int(parts[1])
        except ValueError:
            return OpResult(success=False, message=f"Invalid size values: {size_str!r}")

    ws = ctx.active_sheet
    ws.add_image(img, cell_str.upper())

    return OpResult(success=True, message=f"Image added at {cell_str.upper()}", prefix="+")


def op_link(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Add or remove hyperlinks.

    Syntax:
      link CELL url:URL [text:"TEXT"]
      link CELL sheet:SHEET!CELL [text:"TEXT"]
      link off CELL
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: link CELL url:URL | link off CELL")

    # Handle "link off CELL"
    if op.positionals[0].lower() == "off":
        if len(op.positionals) < 2:
            return OpResult(success=False, message="Usage: link off CELL")
        cell_str = op.positionals[1]
        cell_ref = parse_cell_ref(cell_str)
        if not cell_ref:
            return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

        ws = ctx.active_sheet
        cell = ws.cell(row=cell_ref.row, column=cell_ref.col)
        cell.hyperlink = None
        addr = f"{index_to_col(cell_ref.col)}{cell_ref.row}"
        return OpResult(success=True, message=f"Removed link from {addr}", prefix="-")

    cell_str = op.positionals[0]
    cell_ref = parse_cell_ref(cell_str)
    if not cell_ref:
        return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

    ws = ctx.active_sheet
    cell = ws.cell(row=cell_ref.row, column=cell_ref.col)
    addr = f"{index_to_col(cell_ref.col)}{cell_ref.row}"

    url = op.params.get("url")
    sheet_ref = op.params.get("sheet")
    text = op.params.get("text")

    if url:
        # External hyperlink
        cell.hyperlink = url
        cell.value = text or url
        cell.style = "Hyperlink"
        return OpResult(success=True, message=f"Link added at {addr}: {url}", prefix="+")

    elif sheet_ref:
        # Internal hyperlink: sheet:Sheet2!A1
        # Build internal link
        if "!" in sheet_ref:
            target = f"#{sheet_ref}"
            # Ensure sheet part is quoted if needed
            sheet_part, _, cell_part = sheet_ref.partition("!")
            if sheet_part not in ctx.wb.sheetnames:
                return OpResult(success=False, message=f"Sheet '{sheet_part}' not found")
            target = f"#'{sheet_part}'!{cell_part}"
        else:
            # Just a sheet name — link to A1
            if sheet_ref not in ctx.wb.sheetnames:
                return OpResult(success=False, message=f"Sheet '{sheet_ref}' not found")
            target = f"#'{sheet_ref}'!A1"

        cell.hyperlink = target
        cell.value = text or sheet_ref
        cell.style = "Hyperlink"
        return OpResult(success=True, message=f"Internal link added at {addr}: {target}", prefix="+")

    else:
        return OpResult(success=False, message="Missing 'url' or 'sheet' parameter. Usage: link CELL url:URL")


def op_comment(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Add or remove cell comments.

    Syntax:
      comment CELL "TEXT"
      comment off CELL
    """
    if not op.positionals:
        return OpResult(success=False, message='Usage: comment CELL "TEXT" | comment off CELL')

    # Handle "comment off CELL"
    if op.positionals[0].lower() == "off":
        if len(op.positionals) < 2:
            return OpResult(success=False, message="Usage: comment off CELL")
        cell_str = op.positionals[1]
        cell_ref = parse_cell_ref(cell_str)
        if not cell_ref:
            return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

        ws = ctx.active_sheet
        cell = ws.cell(row=cell_ref.row, column=cell_ref.col)
        cell.comment = None
        addr = f"{index_to_col(cell_ref.col)}{cell_ref.row}"
        return OpResult(success=True, message=f"Removed comment from {addr}", prefix="-")

    # Add comment
    if len(op.positionals) < 2:
        return OpResult(success=False, message='Usage: comment CELL "TEXT"')

    cell_str = op.positionals[0]
    comment_text = op.positionals[1]

    cell_ref = parse_cell_ref(cell_str)
    if not cell_ref:
        return OpResult(success=False, message=f"Invalid cell reference: {cell_str!r}")

    ws = ctx.active_sheet
    cell = ws.cell(row=cell_ref.row, column=cell_ref.col)
    cell.comment = Comment(comment_text, "fcp-sheets")

    addr = f"{index_to_col(cell_ref.col)}{cell_ref.row}"
    return OpResult(success=True, message=f"Comment added at {addr}", prefix="+")


def op_protect(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Protect the active sheet.

    Syntax: protect [password:PWD]
    """
    ws = ctx.active_sheet
    ws.protection.sheet = True

    password = op.params.get("password")
    if password:
        ws.protection.password = password

    msg = f"Sheet '{ws.title}' protected"
    if password:
        msg += " (with password)"
    return OpResult(success=True, message=msg, prefix="*")


def op_unprotect(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Unprotect the active sheet.

    Syntax: unprotect [password:PWD]
    """
    ws = ctx.active_sheet
    ws.protection.sheet = False
    ws.protection._password = None

    return OpResult(success=True, message=f"Sheet '{ws.title}' unprotected", prefix="*")


def op_lock(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Lock cell range.

    Syntax: lock RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: lock RANGE")

    range_str = op.positionals[0]
    ws = ctx.active_sheet
    count = 0

    # Try as range
    range_ref = parse_range_ref(range_str)
    if isinstance(range_ref, RangeRef):
        for row in range(range_ref.start.row, range_ref.end.row + 1):
            for col in range(range_ref.start.col, range_ref.end.col + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=True)
                count += 1
        return OpResult(success=True, message=f"Locked {count} cells in {range_str}", prefix="*")

    # Try as single cell
    cell_ref = parse_cell_ref(range_str)
    if cell_ref:
        ws.cell(row=cell_ref.row, column=cell_ref.col).protection = Protection(locked=True)
        return OpResult(success=True, message=f"Locked {range_str}", prefix="*")

    return OpResult(success=False, message=f"Invalid range: {range_str!r}")


def op_unlock(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Unlock cell range.

    Syntax: unlock RANGE
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: unlock RANGE")

    range_str = op.positionals[0]
    ws = ctx.active_sheet
    count = 0

    # Try as range
    range_ref = parse_range_ref(range_str)
    if isinstance(range_ref, RangeRef):
        for row in range(range_ref.start.row, range_ref.end.row + 1):
            for col in range(range_ref.start.col, range_ref.end.col + 1):
                ws.cell(row=row, column=col).protection = Protection(locked=False)
                count += 1
        return OpResult(success=True, message=f"Unlocked {count} cells in {range_str}", prefix="*")

    # Try as single cell
    cell_ref = parse_cell_ref(range_str)
    if cell_ref:
        ws.cell(row=cell_ref.row, column=cell_ref.col).protection = Protection(locked=False)
        return OpResult(success=True, message=f"Unlocked {range_str}", prefix="*")

    return OpResult(success=False, message=f"Invalid range: {range_str!r}")


# Paper size constants
_PAPER_SIZES = {
    "letter": 1,   # PAPERSIZE_LETTER
    "a4": 9,       # PAPERSIZE_A4
    "legal": 5,    # PAPERSIZE_LEGAL
    "a3": 8,
    "tabloid": 3,
}


def op_page_setup(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Configure page setup for printing.

    Syntax: page-setup [orient:landscape|portrait] [paper:letter|a4|legal]
            [margins:T,R,B,L] [header:TEXT] [footer:TEXT] [print-area:RANGE]
            [print-title-rows:ROW_RANGE] [print-title-cols:COL_RANGE]
            [fit-width:N] [fit-height:N] [gridlines] [center-h] [center-v]
    """
    ws = ctx.active_sheet
    changes = []

    # Orientation
    orient = op.params.get("orient")
    if orient:
        orient_lower = orient.lower()
        if orient_lower not in ("landscape", "portrait"):
            return OpResult(success=False, message=f"Invalid orientation: {orient!r}. Use 'landscape' or 'portrait'")
        ws.page_setup.orientation = orient_lower
        changes.append(f"orientation={orient_lower}")

    # Paper size
    paper = op.params.get("paper")
    if paper:
        paper_lower = paper.lower()
        paper_code = _PAPER_SIZES.get(paper_lower)
        if paper_code is None:
            return OpResult(
                success=False,
                message=f"Unknown paper size: {paper!r}. Use: {', '.join(_PAPER_SIZES.keys())}",
            )
        ws.page_setup.paperSize = paper_code
        changes.append(f"paper={paper_lower}")

    # Margins (T,R,B,L)
    margins_str = op.params.get("margins")
    if margins_str:
        parts = margins_str.split(",")
        if len(parts) != 4:
            return OpResult(success=False, message="Margins must be T,R,B,L (4 values)")
        try:
            top, right, bottom, left = [float(p.strip()) for p in parts]
        except ValueError:
            return OpResult(success=False, message=f"Invalid margin values: {margins_str!r}")
        ws.page_margins = PageMargins(top=top, right=right, bottom=bottom, left=left)
        changes.append("margins set")

    # Header
    header_text = op.params.get("header")
    if header_text:
        ws.oddHeader.center.text = header_text
        changes.append(f"header set")

    # Footer
    footer_text = op.params.get("footer")
    if footer_text:
        ws.oddFooter.center.text = footer_text
        changes.append(f"footer set")

    # Print area
    print_area = op.params.get("print-area")
    if print_area:
        ws.print_area = print_area
        changes.append(f"print-area={print_area}")

    # Print title rows/cols (repeat at top/left of each printed page)
    title_rows = op.params.get("print-title-rows")
    if title_rows:
        ws.print_title_rows = title_rows
        changes.append(f"print-title-rows={ws.print_title_rows}")

    title_cols = op.params.get("print-title-cols")
    if title_cols:
        ws.print_title_cols = title_cols
        changes.append(f"print-title-cols={ws.print_title_cols}")

    # Fit to width/height
    fit_width = op.params.get("fit-width")
    fit_height = op.params.get("fit-height")
    if fit_width:
        try:
            ws.page_setup.fitToWidth = int(fit_width)
        except ValueError:
            return OpResult(success=False, message=f"Invalid fit-width: {fit_width!r}")
        changes.append(f"fit-width={fit_width}")
    if fit_height:
        try:
            ws.page_setup.fitToHeight = int(fit_height)
        except ValueError:
            return OpResult(success=False, message=f"Invalid fit-height: {fit_height!r}")
        changes.append(f"fit-height={fit_height}")

    if fit_width or fit_height:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Gridlines (flag check — appears in positionals for flags)
    if "gridlines" in op.positionals:
        ws.print_options.gridLines = True
        changes.append("gridlines=on")

    # Center horizontal
    if "center-h" in op.positionals:
        ws.print_options.horizontalCentered = True
        changes.append("center-h")

    # Center vertical
    if "center-v" in op.positionals:
        ws.print_options.verticalCentered = True
        changes.append("center-v")

    if not changes:
        return OpResult(success=True, message="No page-setup changes specified", prefix="!")

    return OpResult(success=True, message=f"Page setup: {', '.join(changes)}", prefix="*")


HANDLERS: dict[str, callable] = {
    "name": op_name,
    "image": op_image,
    "link": op_link,
    "comment": op_comment,
    "protect": op_protect,
    "unprotect": op_unprotect,
    "lock": op_lock,
    "unlock": op_unlock,
    "page-setup": op_page_setup,
}
