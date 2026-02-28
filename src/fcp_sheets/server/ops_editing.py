"""Editing operation handlers — copy, move, sort, insert/delete, remove."""

from __future__ import annotations

from copy import copy

from openpyxl.worksheet.worksheet import Worksheet

from fcp_core import OpResult, ParsedOp

from fcp_sheets.model.refs import (
    CellRef,
    RangeRef,
    col_to_index,
    index_to_col,
    parse_cell_ref,
    parse_range_ref,
)
from fcp_sheets.server.resolvers import SheetsOpContext, resolve_range_to_cells


def op_remove(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Clear values in matched cells.

    Syntax: remove RANGE | remove @SELECTOR
    For MVP, accepts a range as positional[0].
    """
    range_str = None
    if op.positionals:
        range_str = op.positionals[0]
    elif op.selectors:
        # For MVP, treat @range:X as range X
        for sel in op.selectors:
            if sel.startswith("range:"):
                range_str = sel[6:]
                break
        if range_str is None:
            return OpResult(success=False, message="Unsupported selector for remove (MVP supports range only)")

    if not range_str:
        return OpResult(success=False, message="Usage: remove RANGE")

    ws = ctx.active_sheet
    count = 0

    # Try as range
    range_ref = parse_range_ref(range_str)
    if isinstance(range_ref, RangeRef):
        for row in range(range_ref.start.row, range_ref.end.row + 1):
            for col in range(range_ref.start.col, range_ref.end.col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.value = None
                    count += 1
        return OpResult(success=True, message=f"Cleared {count} cells in {range_str}", prefix="-")

    # Try as single cell
    cell_ref = parse_cell_ref(range_str)
    if cell_ref:
        cell = ws.cell(row=cell_ref.row, column=cell_ref.col)
        cell.value = None
        return OpResult(success=True, message=f"Cleared {range_str}", prefix="-")

    return OpResult(success=False, message=f"Invalid range: {range_str!r}")


def _copy_cell(src_ws: Worksheet, src_row: int, src_col: int,
               dst_ws: Worksheet, dst_row: int, dst_col: int) -> None:
    """Copy value and formatting from one cell to another."""
    src_cell = src_ws.cell(row=src_row, column=src_col)
    dst_cell = dst_ws.cell(row=dst_row, column=dst_col)

    # Copy value
    dst_cell.value = src_cell.value

    # Copy formatting attributes
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def op_copy(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Copy range to destination.

    Syntax: copy RANGE to:CELL [sheet:NAME]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: copy RANGE to:CELL [sheet:NAME]")

    range_str = op.positionals[0]
    dest_str = op.params.get("to")
    if not dest_str:
        return OpResult(success=False, message="Missing 'to' parameter. Usage: copy RANGE to:CELL")

    # Parse source range
    range_ref = parse_range_ref(range_str)
    if not isinstance(range_ref, RangeRef):
        # Try single cell as 1x1 range
        cell_ref = parse_cell_ref(range_str)
        if cell_ref:
            range_ref = RangeRef(
                start=CellRef(col=cell_ref.col, row=cell_ref.row),
                end=CellRef(col=cell_ref.col, row=cell_ref.row),
            )
        else:
            return OpResult(success=False, message=f"Invalid source range: {range_str!r}")

    # Parse destination cell
    dest_ref = parse_cell_ref(dest_str)
    if not dest_ref:
        return OpResult(success=False, message=f"Invalid destination cell: {dest_str!r}")

    src_ws = ctx.active_sheet

    # Determine destination sheet
    sheet_name = op.params.get("sheet")
    if sheet_name:
        if sheet_name not in ctx.wb.sheetnames:
            return OpResult(success=False, message=f"Sheet '{sheet_name}' not found")
        dst_ws = ctx.wb[sheet_name]
    else:
        dst_ws = src_ws

    # Calculate offsets
    row_offset = dest_ref.row - range_ref.start.row
    col_offset = dest_ref.col - range_ref.start.col

    # Copy each cell
    count = 0
    for row in range(range_ref.start.row, range_ref.end.row + 1):
        for col in range(range_ref.start.col, range_ref.end.col + 1):
            _copy_cell(src_ws, row, col, dst_ws, row + row_offset, col + col_offset)
            count += 1

    # Update index for destination
    end_row = range_ref.end.row + row_offset
    end_col = range_ref.end.col + col_offset
    ctx.index.expand_bounds(dst_ws.title, dest_ref.row, dest_ref.col)
    ctx.index.expand_bounds(dst_ws.title, end_row, end_col)

    dest_label = f"{index_to_col(dest_ref.col)}{dest_ref.row}"
    sheet_info = f" on '{sheet_name}'" if sheet_name else ""
    return OpResult(success=True, message=f"Copied {range_str} to {dest_label}{sheet_info} ({count} cells)", prefix="+")


def op_move(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Move range to destination (copy then clear source).

    Syntax: move RANGE to:CELL [sheet:NAME]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: move RANGE to:CELL [sheet:NAME]")

    range_str = op.positionals[0]
    dest_str = op.params.get("to")
    if not dest_str:
        return OpResult(success=False, message="Missing 'to' parameter. Usage: move RANGE to:CELL")

    # Parse source range
    range_ref = parse_range_ref(range_str)
    if not isinstance(range_ref, RangeRef):
        cell_ref = parse_cell_ref(range_str)
        if cell_ref:
            range_ref = RangeRef(
                start=CellRef(col=cell_ref.col, row=cell_ref.row),
                end=CellRef(col=cell_ref.col, row=cell_ref.row),
            )
        else:
            return OpResult(success=False, message=f"Invalid source range: {range_str!r}")

    # Parse destination cell
    dest_ref = parse_cell_ref(dest_str)
    if not dest_ref:
        return OpResult(success=False, message=f"Invalid destination cell: {dest_str!r}")

    src_ws = ctx.active_sheet

    # Determine destination sheet
    sheet_name = op.params.get("sheet")
    if sheet_name:
        if sheet_name not in ctx.wb.sheetnames:
            return OpResult(success=False, message=f"Sheet '{sheet_name}' not found")
        dst_ws = ctx.wb[sheet_name]
    else:
        dst_ws = src_ws

    # Calculate offsets
    row_offset = dest_ref.row - range_ref.start.row
    col_offset = dest_ref.col - range_ref.start.col

    # Copy each cell first
    count = 0
    for row in range(range_ref.start.row, range_ref.end.row + 1):
        for col in range(range_ref.start.col, range_ref.end.col + 1):
            _copy_cell(src_ws, row, col, dst_ws, row + row_offset, col + col_offset)
            count += 1

    # Clear source cells (only if source != dest overlap area)
    for row in range(range_ref.start.row, range_ref.end.row + 1):
        for col in range(range_ref.start.col, range_ref.end.col + 1):
            dst_row = row + row_offset
            dst_col = col + col_offset
            # Only clear if this source cell doesn't overlap with a destination cell
            if (dst_ws is not src_ws or
                    dst_row < range_ref.start.row or dst_row > range_ref.end.row or
                    dst_col < range_ref.start.col or dst_col > range_ref.end.col or
                    row != dst_row or col != dst_col):
                src_cell = src_ws.cell(row=row, column=col)
                src_cell.value = None

    # Update index
    end_row = range_ref.end.row + row_offset
    end_col = range_ref.end.col + col_offset
    ctx.index.expand_bounds(dst_ws.title, dest_ref.row, dest_ref.col)
    ctx.index.expand_bounds(dst_ws.title, end_row, end_col)

    dest_label = f"{index_to_col(dest_ref.col)}{dest_ref.row}"
    sheet_info = f" on '{sheet_name}'" if sheet_name else ""
    return OpResult(success=True, message=f"Moved {range_str} to {dest_label}{sheet_info} ({count} cells)", prefix="*")


def op_sort(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Sort range by column(s).

    Syntax: sort RANGE by:COL [dir:asc|desc] [by2:COL dir2:asc|desc]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: sort RANGE by:COL [dir:asc|desc]")

    range_str = op.positionals[0]
    by_col = op.params.get("by")
    if not by_col:
        return OpResult(success=False, message="Missing 'by' parameter. Usage: sort RANGE by:COL")

    # Parse direction
    direction = op.params.get("dir", "asc").lower()
    if direction not in ("asc", "desc"):
        return OpResult(success=False, message=f"Invalid direction: {direction!r}. Use 'asc' or 'desc'")

    # Parse range
    range_ref = parse_range_ref(range_str)
    if not isinstance(range_ref, RangeRef):
        return OpResult(success=False, message=f"Invalid range: {range_str!r}")

    # Parse sort column — can be a letter like "B" or a full col ref
    sort_col_idx = col_to_index(by_col.upper()) if by_col.isalpha() else None
    if sort_col_idx is None:
        try:
            sort_col_idx = int(by_col)
        except ValueError:
            return OpResult(success=False, message=f"Invalid sort column: {by_col!r}")

    # Validate sort column is within range
    if sort_col_idx < range_ref.start.col or sort_col_idx > range_ref.end.col:
        return OpResult(
            success=False,
            message=f"Sort column {index_to_col(sort_col_idx)} is outside range {range_str}",
        )

    ws = ctx.active_sheet

    # Read all rows into memory
    rows_data: list[list] = []
    for row in range(range_ref.start.row, range_ref.end.row + 1):
        row_vals = []
        for col in range(range_ref.start.col, range_ref.end.col + 1):
            row_vals.append(ws.cell(row=row, column=col).value)
        rows_data.append(row_vals)

    # Sort key index (relative to range start col)
    key_idx = sort_col_idx - range_ref.start.col
    reverse = direction == "desc"

    # Secondary sort
    by2_col = op.params.get("by2")
    dir2 = op.params.get("dir2", "asc").lower()
    reverse2 = dir2 == "desc"

    if by2_col:
        sort_col2_idx = col_to_index(by2_col.upper()) if by2_col.isalpha() else None
        if sort_col2_idx is None:
            try:
                sort_col2_idx = int(by2_col)
            except ValueError:
                return OpResult(success=False, message=f"Invalid secondary sort column: {by2_col!r}")
        key2_idx = sort_col2_idx - range_ref.start.col

        def sort_key(row):
            v1 = row[key_idx]
            v2 = row[key2_idx]
            # Handle None — sort to end
            k1 = (0, v1) if v1 is not None else (1, "")
            k2 = (0, v2) if v2 is not None else (1, "")
            return k1, k2

        # For multi-key sort with different directions, use two-pass approach
        # Sort by secondary key first (stable sort), then primary
        if reverse != reverse2:
            rows_data.sort(key=lambda r: (0, r[key2_idx]) if r[key2_idx] is not None else (1, ""), reverse=reverse2)
            rows_data.sort(key=lambda r: (0, r[key_idx]) if r[key_idx] is not None else (1, ""), reverse=reverse)
        else:
            rows_data.sort(key=sort_key, reverse=reverse)
    else:
        def sort_key_single(row):
            v = row[key_idx]
            return (0, v) if v is not None else (1, "")

        rows_data.sort(key=sort_key_single, reverse=reverse)

    # Write sorted data back
    for i, row_vals in enumerate(rows_data):
        row = range_ref.start.row + i
        for j, val in enumerate(row_vals):
            col = range_ref.start.col + j
            ws.cell(row=row, column=col).value = val

    num_rows = range_ref.end.row - range_ref.start.row + 1
    sort_col_label = index_to_col(sort_col_idx)
    return OpResult(
        success=True,
        message=f"Sorted {range_str} by column {sort_col_label} ({direction}), {num_rows} rows",
        prefix="*",
    )


def op_insert_row(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Insert rows.

    Syntax: insert-row ROW [count:N]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: insert-row ROW [count:N]")

    try:
        row = int(op.positionals[0])
    except ValueError:
        return OpResult(success=False, message=f"Invalid row number: {op.positionals[0]!r}")

    if row < 1:
        return OpResult(success=False, message="Row number must be >= 1")

    count = 1
    if "count" in op.params:
        try:
            count = int(op.params["count"])
        except ValueError:
            return OpResult(success=False, message=f"Invalid count: {op.params['count']!r}")
        if count < 1:
            return OpResult(success=False, message="Count must be >= 1")

    ws = ctx.active_sheet
    ws.insert_rows(row, count)

    plural = "s" if count > 1 else ""
    return OpResult(success=True, message=f"Inserted {count} row{plural} at row {row}", prefix="+")


def op_insert_col(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Insert columns.

    Syntax: insert-col COL [count:N]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: insert-col COL [count:N]")

    col_str = op.positionals[0]

    # Accept column letter or number
    if col_str.isalpha():
        col_idx = col_to_index(col_str.upper())
    else:
        try:
            col_idx = int(col_str)
        except ValueError:
            return OpResult(success=False, message=f"Invalid column: {col_str!r}")

    if col_idx < 1:
        return OpResult(success=False, message="Column index must be >= 1")

    count = 1
    if "count" in op.params:
        try:
            count = int(op.params["count"])
        except ValueError:
            return OpResult(success=False, message=f"Invalid count: {op.params['count']!r}")
        if count < 1:
            return OpResult(success=False, message="Count must be >= 1")

    ws = ctx.active_sheet
    ws.insert_cols(col_idx, count)

    col_label = index_to_col(col_idx)
    plural = "s" if count > 1 else ""
    return OpResult(success=True, message=f"Inserted {count} column{plural} at column {col_label}", prefix="+")


def op_delete_row(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Delete rows.

    Syntax: delete-row ROW [count:N]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: delete-row ROW [count:N]")

    try:
        row = int(op.positionals[0])
    except ValueError:
        return OpResult(success=False, message=f"Invalid row number: {op.positionals[0]!r}")

    if row < 1:
        return OpResult(success=False, message="Row number must be >= 1")

    count = 1
    if "count" in op.params:
        try:
            count = int(op.params["count"])
        except ValueError:
            return OpResult(success=False, message=f"Invalid count: {op.params['count']!r}")
        if count < 1:
            return OpResult(success=False, message="Count must be >= 1")

    ws = ctx.active_sheet
    ws.delete_rows(row, count)

    plural = "s" if count > 1 else ""
    return OpResult(success=True, message=f"Deleted {count} row{plural} at row {row}", prefix="-")


def op_delete_col(op: ParsedOp, ctx: SheetsOpContext) -> OpResult:
    """Delete columns.

    Syntax: delete-col COL [count:N]
    """
    if not op.positionals:
        return OpResult(success=False, message="Usage: delete-col COL [count:N]")

    col_str = op.positionals[0]

    if col_str.isalpha():
        col_idx = col_to_index(col_str.upper())
    else:
        try:
            col_idx = int(col_str)
        except ValueError:
            return OpResult(success=False, message=f"Invalid column: {col_str!r}")

    if col_idx < 1:
        return OpResult(success=False, message="Column index must be >= 1")

    count = 1
    if "count" in op.params:
        try:
            count = int(op.params["count"])
        except ValueError:
            return OpResult(success=False, message=f"Invalid count: {op.params['count']!r}")
        if count < 1:
            return OpResult(success=False, message="Count must be >= 1")

    ws = ctx.active_sheet
    ws.delete_cols(col_idx, count)

    col_label = index_to_col(col_idx)
    plural = "s" if count > 1 else ""
    return OpResult(success=True, message=f"Deleted {count} column{plural} at column {col_label}", prefix="-")


HANDLERS: dict[str, callable] = {
    "remove": op_remove,
    "copy": op_copy,
    "move": op_move,
    "sort": op_sort,
    "insert-row": op_insert_row,
    "insert-col": op_insert_col,
    "delete-row": op_delete_row,
    "delete-col": op_delete_col,
}
