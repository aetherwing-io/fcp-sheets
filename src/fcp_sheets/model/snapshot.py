"""Byte-snapshot undo/redo for openpyxl workbooks.

SnapshotEvent captures before/after states as bytes from
wb.save(BytesIO) / load_workbook(BytesIO). This enables
simple, correct undo/redo for all workbook mutations.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook, load_workbook


@dataclass
class SnapshotEvent:
    """Event type for byte-snapshot undo/redo."""

    type: str = "snapshot"
    before: bytes = field(default=b"", repr=False)
    after: bytes = field(default=b"", repr=False)
    summary: str = ""


def snapshot_workbook(wb: Workbook) -> bytes:
    """Serialize a workbook to bytes via BytesIO."""
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def restore_workbook(data: bytes) -> Workbook:
    """Deserialize a workbook from snapshot bytes."""
    buf = BytesIO(data)
    return load_workbook(buf)


class SheetsModel:
    """Thin wrapper around openpyxl Workbook for in-place undo/redo.

    The session dispatcher holds a reference to this object.
    reverse_event/replay_event replace self.wb in place so
    the session reference stays valid.
    """

    def __init__(self, title: str = "Untitled", wb: Workbook | None = None):
        self.title = title
        self.wb: Workbook = wb or Workbook()
        self.file_path: str | None = None

    def snapshot(self) -> bytes:
        """Take a byte snapshot of the current workbook state."""
        return snapshot_workbook(self.wb)

    def restore(self, data: bytes) -> None:
        """Replace the workbook from snapshot bytes (in-place for undo/redo)."""
        self.wb = restore_workbook(data)
